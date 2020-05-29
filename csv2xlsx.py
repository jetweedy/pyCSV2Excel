# -----------------------------------------------------------------------
# pip install xlsxwriter, openpyxl, pandas
# -----------------------------------------------------------------------

#  python3 csv2xlsx.py targetfile.xlsx sourcefile1 sourcefile2 --tabnames"Tab 1|Tab B"

import os
import csv
import sys
import glob
import time;
import argparse;

ts1 = time.time();

import pandas as pd
from pandas.io.excel import ExcelWriter
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

#print('Number of arguments:', len(sys.argv), 'arguments.')
#print('Argument List:', str(sys.argv))

# https://stackoverflow.com/questions/40001892/reading-named-command-arguments
parser=argparse.ArgumentParser()
parser.add_argument('--tabnames', help='tab names separated by pipe')
#args=parser.parse_args()
args, unknown = parser.parse_known_args()
print("\nargs:", args);
if (args.tabnames != None):
    print("args.tabnames: ", args.tabnames);
print("\nsys.argv", sys.argv);

if (len(sys.argv) > 1):

    all_args = sys.argv;
    all_files = [];
    for a in all_args:
        if (a[0] != "-"):
            all_files.append(a);
    print("\nall_files: ", all_files);
    
    all_files.pop(0);
    targetFile = all_files.pop(0);

    #all_files = ["csv1.csv", "csv5.csv"];
    #all_files = ["csv3.csv", "csv4.csv"];
    
    tabnames = [];
    if (args.tabnames != None):
        tabnames = args.tabnames.split("|");
    print("\ntabnames: ", tabnames);

    df_from_each_file = (pd.read_csv(f, low_memory=False) for f in all_files)
    writer = pd.ExcelWriter(targetFile, engine='xlsxwriter');
    for idx, df in enumerate(df_from_each_file):
        if (len(tabnames)>idx):
            print(tabnames[idx]);
            df.to_excel(writer, sheet_name=tabnames[idx], index=False, encoding='utf-8')
        else:
            print("Sheet {0}");
            df.to_excel(writer, sheet_name='Sheet {0}'.format(idx), index=False, encoding='utf-8')
    writer.save()

ts2 = time.time();
print("\nExecution time: ", ts2-ts1);


### Below is code demonstrating how to do some more dynamic cell creation
"""
book = load_workbook('result.xlsx')
for i in range(2, book["Sheet 0"].max_row+1):
#    print(".");
    book["Sheet 0"]['C'+str(i)] = i
    book["Sheet 0"]['D'+str(i)] = i+1
    book["Sheet 0"]['E'+str(i)] = '=SUM(C'+str(i)+',D'+str(i)+')'

dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)
book["Sheet 0"].add_data_validation(dv)
dv.add('F2:F'+str(book["Sheet 0"].max_row+1));
book["Sheet 0"].column_dimensions['F'].width = 100
book.save('result.xlsx')
"""

